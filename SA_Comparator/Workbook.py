import os
import openpyxl
from Converter import *
from openpyxl.styles import PatternFill, Font, Color

current_path = os.getcwd()

def Create_Sheet_Type1(results,table,header=[]):
    
    # Create a new Excel workbook or open an existing one
    try:
        #workbook = openpyxl.load_workbook(os.path.join(current_path, 'Result.xlsx'))
        workbook = openpyxl.load_workbook('D:\SA Comparator\Result.xlsx')
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.active.title = table
    
    # Create or get the desired worksheet based on the sheet number
    
    if table in workbook.sheetnames:
        sheet = workbook[table]
    else:
        sheet = workbook.create_sheet(table)
   
    Create_Header(sheet,header)
    
    Partition_Fill('Data Changed',sheet)
    Text_Fill(results[0],sheet)
    Partition_Fill('Data Removed',sheet)
    Text_Fill(results[1],sheet)
    Partition_Fill('Data Added',sheet)
    Text_Fill(results[2],sheet)

    Adjust_Column(sheet)

    # Save the workbook to a file
    #workbook.save(os.path.join(current_path, 'Result.xlsx'))
    workbook.save('D:\SA Comparator\Result.xlsx')

def Text_Fill(results,sheet):
    
    if len(results) == 0:
        sheet.cell(row=sheet.max_row + 2, column=1).value = "None!!"
        return

    for no,data in enumerate(results, start=1):
        row_index = sheet.max_row + 1
        if no%2 == 0:
            for i in range(max(len(results[no-2]),len(data))):
                    row_color = sheet.cell(row=row_index, column=i+1)
                    row_color.fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
        for col_idx, value in enumerate(data, start=1):
            cell = sheet.cell(row=row_index, column=col_idx)
            cell.value = value
            if no%2 == 0 and results[no-2][0] != data[0]:               
                for i, (elem1, elem2) in enumerate(zip(results[no-2], data)):                 
                    if elem1 != elem2 and i > 1: 
                        cell = sheet.cell(row=row_index, column=i+1)
                        cell.font = Font(color=Color(rgb="FF0000"))

def Partition_Fill(text,sheet):
    sheet.cell(row=sheet.max_row + 2, column=1).value = text
    sheet.cell(row=sheet.max_row, column=1).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

def Create_Sheet_Type2(results,table,header=[]):
    
    # Load the existing workbook
    #workbook = openpyxl.load_workbook(os.path.join(current_path, 'Result.xlsx'))
    workbook = openpyxl.load_workbook('D:\SA Comparator\Result.xlsx')

    # Create a new sheet
    sheet = workbook.create_sheet(title=table)
    Create_Header(sheet,header)

    Partition_Fill('Data missmatch',sheet)

    # Fill Data missmatch
    for no,data in enumerate(results, start=1):
        row_index = sheet.max_row + 1
        if no%2 == 0:
            for i in range(max(len(results[no-2]),len(data))):
                    row_color = sheet.cell(row=row_index, column=i+1)
                    row_color.fill = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
        for col_idx, value in enumerate(data, start=1):
            cell = sheet.cell(row=row_index, column=col_idx)
            cell.value = value
    
    if len(results) == 0:
        sheet.cell(row=sheet.max_row + 2, column=1).value = "None!!"

    Adjust_Column(sheet)
    
    #workbook.save(os.path.join(current_path, 'Result.xlsx'))
    workbook.save('D:\SA Comparator\Result.xlsx')

def Adjust_Column(sheet):
    # Adjusting column width to fit content
    for column in sheet.columns:
        max_length = 0
        column_name = column[0].column_letter  # Get the column name (e.g., 'A', 'B', etc.)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass

        adjusted_width = (max_length + 2) * 1.2  # Adding a little extra width for padding
        sheet.column_dimensions[column_name].width = adjusted_width

def Create_Header(sheet,header):
    # Fill header
 
    for col_idx,head in enumerate(header,start=1):     
        cell = sheet.cell(row=1, column=col_idx)
        cell.fill = PatternFill(start_color="3399FF", end_color="3399FF", fill_type="solid")
        cell.value = head